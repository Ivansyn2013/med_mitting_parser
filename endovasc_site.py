from bs4 import BeautifulSoup as bs
import requests
import openpyxl
import re


def change_col_size(sheet):

    ''' функция принимает лист и выставляет значение ширины колонки по максимальному значения
    величины содержимого ячейки
    '''
    for i in range(1, sheet.max_column + 1):
        cell_obj = [x for x in sheet.columns]
        max_lenth_obj = max(cell_obj, key=lambda x: list(map(lambda z: len(str(z.value)), x)))
        letter = openpyxl.utils.cell.get_column_letter(i)
        sheet.column_dimensions[letter].width = len(max_lenth_obj[0].value)
    return None

URL_TEMPLATE = 'https://endovascular.ru/events/preview'
wb = openpyxl.Workbook()
wb.create_sheet(title='First list', index=0)
work_sh = wb['First list']

headers = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS x 10_10_1) AppleWebkit/537.36) (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36'}

r = requests.get(URL_TEMPLATE, headers=headers)

print("\033[31m {}".format('Status of connection:' + str(r.status_code)))
print("\033[0m")

soup = bs(r.text, 'html.parser')

mit_info = soup.find('div', {'class': 'main-sections__center-blocks row'})

mit_intro = soup.findAll('div', class_='main-sections__block')
some_text = mit_intro[0].text

for em, item in enumerate(mit_intro):

    text = item.text
    res = [x.strip() for x in (re.split('\n|\r', text)) if x != '']
    print(res)
    for idx, parag in enumerate(res):
        cell = work_sh.cell(row=em + 1, column=idx + 1)
        cell.value = parag

change_col_size(work_sh)
wb.save('output/end.xlsx')


